import sys
from openpyxl import Workbook
from openpyxl.styles import Font


def create_multiplication_table(ws, N):
    N = N + 1

    for i in range(2, N+1):
        ws.cell(row=1, column=i, value=i-1).font = Font(bold=True)
        ws.cell(row=i, column=1, value=i-1).font = Font(bold=True)

    #掛け算表を作成
    for i in range(2, N + 1):
        for j in range(1, N):
            ws.cell(row=i, column=j+1, value=f"=({i}-1)*{j}")


def main(N):
    # Excel ワークブックの作成
    wb = Workbook()

    # A = N シートを作成
    ws_A_N = wb.active
    ws_A_N.title = f"multiple{N}"
    create_multiplication_table(ws_A_N, N)

    # A = N+1 シートを作成
    ws_A_N_plus_1 = wb.create_sheet(title=f"multiple{N + 1}")
    create_multiplication_table(ws_A_N_plus_1, N + 1)

    # A = N+2 シートを作成
    ws_A_N_plus_2 = wb.create_sheet(title=f"multiple{N + 2}")
    create_multiplication_table(ws_A_N_plus_2, N + 2)

    # Excel ファイルとして保存
    wb.save("multiplicationTable.xlsx")
    print(f"{N}, {N + 1}, {N + 2} の掛け算表が 'multiplicationTable_separateSheets.xlsx' として保存されました。")

if __name__ == "__main__":
    # コマンドライン引数から数字 N を取得
    N = int(sys.argv[1])
    main(N)
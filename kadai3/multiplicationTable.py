import sys
from openpyxl import Workbook

# コマンドライン引数から数字 N を取得
N = int(sys.argv[1])
N = N + 1

# Excel ワークブックの作成
wb = Workbook()
ws = wb.active
ws.title = "Multiplication Table"


for i in range(2, N+1):
    ws.cell(row=1, column=i, value=i-1)
    ws.cell(row=i, column=1, value=i-1)

#掛け算表を作成
for i in range(2, N + 1):
    for j in range(1, N):
        ws.cell(row=i, column=j+1, value=(i-1)*j)

# Excel ファイルとして保存
wb.save("multiplicationTable.xlsx")
